import openpyxl, os

''' a. Rename the sheet '''
os.chdir('D:\Python\CSV\Exercise2')     # Open folder
# Open excel file
workbook = openpyxl.load_workbook('Questions_Poll.xlsx')    
if 'Sheet1' not in workbook.sheetnames:
    Q_sheet = workbook['Sheet1']  # Get Sheet1 from wb
else:
    Q_sheet = workbook[workbook.sheetnames[0]]
Q_sheet.title = 'Questions'   # Change title of Sheet1 -> "Questions"
workbook.save('Questions_Poll.xlsx')    # Save file


''' b. Add a sheet "Results"
    c. Add sheet "QuestionID"
    d. Add sheet "Domains"
    e. Add sheet "Projects"
'''
os.chdir('D:\Python\CSV\Exercise2')     # Open folder
# Open excel file
workbook = openpyxl.load_workbook('Questions_Poll.xlsx')
# Create new sheet
if 'Results' not in workbook.sheetnames:
    results = workbook.create_sheet(index = 1, title = 'Results')   
if 'QuestionID' not in workbook.sheetnames:
    questionid = workbook.create_sheet(index = 2, title = 'QuestionID')
if 'Domains' not in workbook.sheetnames:
    domains = workbook.create_sheet(index = 3, title = 'Domains')
if 'Projects' not in workbook.sheetnames:
    projects = workbook.create_sheet(index = 4, title = 'Projects')

    
# Process data from "Questions" sheet   -> Q_sheet
# & Complete "Results" sheet            -> A_sheet
Q_sheet = workbook['Questions']
A_sheet = workbook['Results']
QID_sheet = workbook['QuestionID']
domain_sheet = workbook['Domains']
proj_sheet = workbook['Projects']

A_sheet['A1'] = 'QuestionID'
A_sheet['B1'] = 'Answer'
A_sheet['C1'] = 'DomainID'
A_sheet['D1'] = 'ProjectID'

QID_sheet['A1'] = 'QuestionID'
QID_sheet['B1'] = 'QuestionText'

domain_sheet['A1'] = 'DomainID'
domain_sheet['B1'] = 'Domain'

proj_sheet['A1'] = 'ProjectID'
proj_sheet['B1'] = 'Project'


# Fill QuestionID & Fill Answer to "Results" sheet
for col in range(1, Q_sheet.max_column + 1):
    count_answer = 0
    row_domain = 2
    row_proj = 2
    Separate_Domain = []
    Separate_Proj = []
    for r in range(2, Q_sheet.max_row + 1):
        # Count the Answers of each Question
        if (Q_sheet.cell(row = r, column = col).value != None) and \
           ('Answer' in Q_sheet.cell(row = r, column = col).value):
            count_answer += 1

        # Search & Fill the Domain
        elif (Q_sheet.cell(row = r, column = col).value != None) and \
             ('Domain' in Q_sheet.cell(row = r, column = col).value):
            
            # Check for duplicate "Domain" values
            if Q_sheet.cell(row = r, column = col).value not in Separate_Domain:
                Temp = Q_sheet.cell(row = r, column = col).value
                Separate_Domain.append(Temp)
                A_sheet.cell(row = row_domain, column = 3).value = 'D' + str(row_domain - 1)
                #A_sheet.cell(row = row_domain, column = 3).value = Temp
                domain_sheet.cell(row = row_domain, column = 1).value = 'D' + str(row_domain - 1)
                domain_sheet.cell(row = row_domain, column = 2).value = Temp
                row_domain += 1

        # Search & Fill the "Project" Sheet
        elif (Q_sheet.cell(row = r, column = col).value != None) and \
             ('Project' in Q_sheet.cell(row = r, column = col).value):
            
            # Check for duplicate "Domain" values
            if Q_sheet.cell(row = r, column = col).value not in Separate_Proj:
                Temp = Q_sheet.cell(row = r, column = col).value
                Separate_Proj.append(Temp)
                A_sheet.cell(row = row_proj, column = 4).value = 'P' + str(row_proj - 1)
                #A_sheet.cell(row = row_proj, column = 4).value = Temp
                proj_sheet.cell(row = row_proj, column = 1).value = 'P' + str(row_proj - 1)
                proj_sheet.cell(row = row_proj, column = 2).value = Temp
                row_proj += 1

    # Fill the Question & Answer
    if 'Question' in Q_sheet.cell(row = 1, column = col).value:
        #A_sheet.cell(row = col+1, column = 1).value = Q_sheet.cell(row = 1, column = col).value
        A_sheet.cell(row = col+1, column = 1).value = 'Q' + str(col)
        A_sheet.cell(row = col+1, column = 2).value = count_answer
        QID_sheet.cell(row = col+1, column = 1).value = 'Q' + str(col)
        QID_sheet.cell(row = col+1, column = 2).value = Q_sheet.cell(row = 1, column = col).value

workbook.save('Questions_Poll.xlsx')    # Save file


''' c. Add sheet "Questions" '''

